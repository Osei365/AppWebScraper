#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import requests
import xlsxwriter
import openpyxl
import os
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException, TimeoutException



# Defining all the xpaths for the fields to be scraped
Name_xpath = "//h1[@class='appx-page-header-2_title']"
Pricing_xpath= "//span[@id='appxListingDetailPageId:AppxLayout:planList:0:planCharges']"
Categories_xpath = "//a[@id='appxListingDetailPageId:AppxLayout:listingCategories:0:firstCat']//strong"
Ratings_xpath = "//span[@id='appxListingDetailPageId:AppxLayout:j_id840:j_id841:j_id844']"
Latest_release_xpath = "//div[@id='AppxListingDetailOverviewTab:listingDetailOverviewTab:appxListingDetailOverviewTabComp:j_id154']//div[@class='appx-extended-detail-subsection-description']"
Tool_intro_information_xpath = ["//div[@class='appx-detail-section-description appx-multi-line-to-fix appx-multi-line-fixed']//div[@class='appx-headline-details-tagline']",
"//div[@class='appx-detail-section-description appx-multi-line-to-fix appx-multi-line-fixed']//p"]
Highlights_xpath = ["//span[@class='appx-highlights-text']",
             "//span[@class='appx-highlights-text']//span[@id='AppxListingDetailOverviewTab:listingDetailOverviewTab:appxListingDetailOverviewTabComp:ftListInventory:{}:featureItem']"]
Description_xpath = "//div[@class='appx-extended-detail-description appx-multi-line-to-fix appx-multi-line-fixed']"
Requirements_xpath = ["//div[@id='AppxListingDetailOverviewTab:listingDetailOverviewTab:appxListingDetailOverviewTabComp:j_id114']//div[@class='appx-extended-detail-subsection-label']",
"//div[@id='AppxListingDetailOverviewTab:listingDetailOverviewTab:appxListingDetailOverviewTabComp:j_id114']//div[@class='appx-extended-detail-subsection-description appx-multi-line-to-fix appx-multi-line-fixed']"]

Additional_info_xpath = ["//div[@id='AppxListingDetailOverviewTab:listingDetailOverviewTab:appxListingDetailOverviewTabComp:j_id149']",
 "//div[@id='AppxListingDetailOverviewTab:listingDetailOverviewTab:appxListingDetailOverviewTabComp:j_id146']",
 "//div[@id='AppxListingDetailOverviewTab:listingDetailOverviewTab:appxListingDetailOverviewTabComp:j_id152']//div[@class='appx-extended-detail-subsection-description']",
 "//div[@id='AppxListingDetailOverviewTab:listingDetailOverviewTab:appxListingDetailOverviewTabComp:j_id154']//div[@class='appx-extended-detail-subsection-description']"]
About_company_xpath = "//div[@class='appx-extended-detail-subsection-segment-double appx-multi-line-to-fix appx-multi-line-fixed']//p[@class='appx-extended-detail-company-description']"
website_xpath = "//div[@class='appx-extended-detail-subsection-description slds-truncate']//a[@data-event='listing-publisher-website']"
email_xpath = "//div[@class='appx-extended-detail-subsection-description slds-truncate']//a[@data-event='listing-publisher-email']"
Address_xpath = "//div[@id='AppxListingDetailOverviewTab:listingDetailOverviewTab:appxListingDetailOverviewTabComp:j_id363']//div[@class='appx-extended-detail-subsection-description']"

def create_workbook(filename, sheetname):
    headers = ['Name', 'Pricing', 'Categories', 'Ratings', 
               'Latest Release', 'Tool intro information', 
               'Highlights', 'Description', 'Requirements', 
               'Additional information', 'About company'
               'Website', 'Email', 'Address']

    workbook = xlsxwriter.Workbook(filename)
    sheet = workbook.add_worksheet(sheetname)
    cols = 'A B C D E F G H I J K L M N'.split()
    for col, header in zip(cols, headers):
        sheet.write('{}1'.format(col), header)

    workbook.close()

def AppScraper(filename, sheetname, URL):
    
    # Load Excel Workbook to save data from website
    xfile = openpyxl.load_workbook(filename)
    sheet = xfile[sheetname]

    # Initiate Selenium Driver for Project
    driver = webdriver.Chrome('chromedriver')
    driver.get(URL)
    top_categories = driver.find_elements_by_xpath("//ul[@class='appx-categories-top-ul']//a")

    # Iterate through top categories to get all apps for each category
    for top_category in top_categories:
        top_cat_link = top_category.get_attribute('href')

        # Make a request for the category in question
        driver.get(top_cat_link)    
        WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.ID, 'appx-load-more-button-id'))) 

        # Each category has a 'show more' button to load more apps on the same page
        # we load all the apps available until the 'show more' doesnt appear again
        while True:

            time.sleep(10)
            element = driver.find_element_by_id('appx-load-more-button-id')
            print(element.text)
            if element.text == 'Show More':
                driver.execute_script("arguments[0].click()", element)
                print('clicked')
            else:
                break



        # Getting the element ID to all the apps on the page
        links = driver.find_elements_by_xpath("//li//a[@data-listing-type='App']")
        app_id = [link.get_attribute('data-listing-id') for link in links]


        # Opening the page of each App to save information or data about the app
        for ID in app_id:
            n = sheet.max_row
            element = driver.find_element_by_xpath("//li//a[@data-listing-id='{}']".format(ID))
            driver.execute_script("arguments[0].click()", element)

    
            try:
                sheet['A{}'.format(str(n+1))] = driver.find_element_by_xpath(Name_xpath).text
                print(driver.find_element_by_xpath(Name_xpath).text)
            except NoSuchElementException:
                sheet['A{}'.format(str(n+1))] = 'Not Indicated'


            try:
                sheet['B{}'.format(str(n+1))] = driver.find_element_by_xpath(Pricing_xpath).text
                print(driver.find_element_by_xpath(Pricing_xpath).text)
            except NoSuchElementException:
                sheet['B{}'.format(str(n+1))] = 'Not Indicated'

            try:
                sheet['C{}'.format(str(n+1))] = driver.find_element_by_xpath(Categories_xpath).text
                print(driver.find_element_by_xpath(Categories_xpath).text)
            except NoSuchElementException:
                sheet['C{}'.format(str(n+1))] = 'Not Indicated'

            try:
                sheet['D{}'.format(str(n+1))] = driver.find_element_by_xpath(Ratings_xpath).get_attribute('class')
                print(driver.find_element_by_xpath(Ratings_xpath).get_attribute('class'))
            except NoSuchElementException:
                sheet['D{}'.format(str(n+1))] = 'Not Indicated'
            try:
                sheet['E{}'.format(str(n+1))] = driver.find_element_by_xpath(Latest_release_xpath).text
                print(driver.find_element_by_xpath(Latest_release_xpath).text)
            except NoSuchElementException:
                sheet['E{}'.format(str(n+1))] = 'Not Indicated'

            try: 

                intro = [driver.find_element_by_xpath(info).text for info in Tool_intro_information_xpath]
                sheet['F{}'.format(str(n+1))] = ' '.join(intro)
                print(' '.join(intro))
            except NoSuchElementException:
                sheet['F{}'.format(str(n+1))] = 'Not Indicated'

            try:
                hls= driver.find_elements_by_xpath(Highlights_xpath[0])
                print(len(hls))
                all_highlight=[]
                for n in range(len(hls)):
                    print(driver.find_element_by_xpath(Highlights_xpath[1].format(str(n))).text)
                    all_highlight.append(driver.find_element_by_xpath(Highlights[1].format(str(n))).text) 
                sheet['G{}'.format(str(n+1))] = ' '.join(all_highlight)
                print(' '.join(all_highlight))
            except NoSuchElementException:
                sheet['G{}'.format(str(n+1))] = 'Not Indicated'

            try:
                element2 = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, Description_xpath)))
                sheet['H{}'.format(str(n+1))] = element2.text
                print(element2.text)
            except NoSuchElementException:
                sheet['H{}'.format(str(n+1))] = 'Not Indicated'

            try:
                requirements_list= [driver.find_element_by_xpath(xpath).text for xpath in Requirements_xpath]
                sheet['I{}'.format(str(n+1))] = ' '.join(requirements_list)
                print(' '.join(requirements_list))
            except NoSuchElementException:
                sheet['I{}'.format(str(n+1))] = 'Not Indicated'


            try:
                Add_info= [driver.find_element_by_xpath(xpath).text for xpath in Additional_info_xpath]
                sheet['J{}'.format(str(n+1))] = ' '.join(Add_info)
                print(' '.join(Add_info))
            except NoSuchElementException:
                sheet['J{}'.format(str(n+1))] = 'Not Indicated'

            try:
                sheet['K{}'.format(str(n+1))] = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, About_company_xpath))).text
                print(WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, About_company_xpath))).text)
            except NoSuchElementException:
                sheet['K{}'.format(str(n+1))] = 'Not Indicated'

            try:
                sheet['L{}'.format(str(n+1))] = driver.find_element_by_xpath(website_xpath).text
                print(driver.find_element_by_xpath(website_xpath).text)
            except NoSuchElementException:
                sheet['L{}'.format(str(n+1))] = 'Not Indicated'

            try:
                sheet['M{}'.format(str(n+1))] = driver.find_element_by_xpath(email_xpath).text
                print(driver.find_element_by_xpath(email_xpath).text)
            except NoSuchElementException:
                sheet['M{}'.format(str(n+1))] = 'Not Indicated'

            try:
                sheet['N{}'.format(cols[13], str(n+1))] = driver.find_element_by_xpath(Address_xpath).text
                print(driver.find_element_by_xpath(Address_xpath).text)
            except NoSuchElementException:
                sheet['N{}'.format(str(n+1))] = 'Not Indicated'

            print('Row added successfully.....')
        

            xfile.save('apps.xlsx')

            driver.back()

            time.sleep(3)
        Print('All apps in {} Category Loaded'.format(top_cat_link))
        xfile.save('apps.xlsx')

 

filename = 'apps.xlsx'        
sheetname = 'Sheet1'    
URL = 'https://appexchange.salesforce.com/'


create_workbook(filename, sheetname)
AppScraper(filename, sheetname, URL)    

